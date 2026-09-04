"""Einmaliger Import des Bundeslebensmittelschluessels 4.0 in tracker.db.

Nicht Teil der Anwendung. Liest quellen/BLS_4_0_Daten_2025_DE.xlsx und befuellt
die Tabellen lebensmittel, naehrstoff und naehrwert.

Grundregel: Unbekannt ist nicht null. Enthaelt eine Wertzelle einen Strich, "TR",
eine Angabe unterhalb der Nachweis- oder Bestimmungsgrenze oder ist sie leer, wird
KEINE Zeile in naehrwert angelegt. Eine echte 0 aus der Quelle wird gespeichert.

Aufruf aus dem Projektordner:
    python importe/import_bls.py
    python importe/import_bls.py --ersetzen   # vorhandene Eintraege aktualisieren
"""

import argparse
import sqlite3
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl fehlt. Installieren mit: pip install openpyxl")


# Das Skript liegt in importe/, die Datenbank und die Datenordner eine Ebene darueber.
BASIS = Path(__file__).resolve().parent.parent
QUELLE = BASIS / "quellen" / "BLS_4_0_Daten_2025_DE.xlsx"
DATENBANK = BASIS / "tracker.db"

# Spalten der Quelldatei, 1-basiert wie in Excel
SPALTE_CODE = 1
SPALTE_BEZEICHNUNG = 2
ERWARTETE_SPALTENZAHL = 418

# Naehrstoffe, die uebernommen werden.
# id, bls_spalte, name, einheit, gruppe, uebergeordnet (bls_spalte), wertspalte, herkunftsspalte
NAEHRSTOFFE = [
    (1, "ENERCC", "Energie", "kcal", "energie", None, 7, 8),
    (2, "PROT625", "Protein", "g", "makronaehrstoff", None, 13, 14),
    (3, "FAT", "Fett", "g", "makronaehrstoff", None, 16, 17),
    (4, "FASAT", "Gesättigte Fettsäuren", "g", "fett", "FAT", 247, 248),
    (5, "CHO", "Kohlenhydrate", "g", "makronaehrstoff", None, 19, 20),
    (6, "SUGAR", "Zucker", "g", "kohlenhydrat", "CHO", 220, 221),
    (7, "LACS", "Lactose", "g", "kohlenhydrat", "SUGAR", 217, 218),
    # Salz wird im BLS als Natrium mal 2,5 berechnet. NA ist damit uebergeordnet,
    # sonst gingen beide gemeinsam in eine Summe ein.
    (8, "NACL", "Salz", "g", "mineralstoff", "NA", 121, 122),
    # Vitamine. Es werden ausschliesslich die Summenparameter uebernommen, nicht
    # ihre Einzelkomponenten: VITA statt Retinol und Carotinoiden, VITD statt
    # Ergo- und Cholecalciferol, VITK statt K1 und K2, FOL statt Folat und
    # Folsaeure, NIAEQ statt Niacin und Tryptophananteil. Deshalb steht bei
    # diesen fuenf kein uebergeordnet_id: ihre Komponenten fehlen in der Tabelle.
    (9, "VITA", "Vitamin A, Retinol-Äquivalent", "µg", "vitamin", None, 34, 35),
    (10, "VITD", "Vitamin D", "µg", "vitamin", None, 49, 50),
    (11, "VITE", "Vitamin E", "mg", "vitamin", None, 58, 59),
    (12, "VITK", "Vitamin K", "µg", "vitamin", None, 76, 77),
    (13, "THIA", "Vitamin B1", "mg", "vitamin", None, 85, 86),
    (14, "RIBF", "Vitamin B2", "mg", "vitamin", None, 88, 89),
    (15, "NIAEQ", "Niacin-Äquivalent", "mg", "vitamin", None, 91, 92),
    (16, "PANTAC", "Pantothensäure", "mg", "vitamin", None, 97, 98),
    (17, "VITB6", "Vitamin B6", "µg", "vitamin", None, 100, 101),
    (18, "BIOT", "Biotin", "µg", "vitamin", None, 103, 104),
    (19, "FOL", "Folat-Äquivalent", "µg", "vitamin", None, 106, 107),
    (20, "VITB12", "Vitamin B12", "µg", "vitamin", None, 115, 116),
    (21, "VITC", "Vitamin C", "mg", "vitamin", None, 118, 119),
    # Mengenelemente
    (22, "NA", "Natrium", "mg", "mineralstoff", None, 124, 125),
    (23, "CLD", "Chlorid", "mg", "mineralstoff", None, 127, 128),
    (24, "K", "Kalium", "mg", "mineralstoff", None, 130, 131),
    (25, "CA", "Calcium", "mg", "mineralstoff", None, 133, 134),
    (26, "MG", "Magnesium", "mg", "mineralstoff", None, 136, 137),
    (27, "P", "Phosphor", "mg", "mineralstoff", None, 139, 140),
    # Spurenelemente
    (28, "FE", "Eisen", "mg", "spurenelement", None, 145, 146),
    (29, "ZN", "Zink", "mg", "spurenelement", None, 148, 149),
    (30, "ID", "Iodid", "µg", "spurenelement", None, 151, 152),
    (31, "CU", "Kupfer", "µg", "spurenelement", None, 154, 155),
    (32, "MN", "Mangan", "µg", "spurenelement", None, 157, 158),
    (33, "FD", "Fluorid", "µg", "spurenelement", None, 160, 161),
    (34, "CR", "Chrom", "µg", "spurenelement", None, 163, 164),
    (35, "MO", "Molybdän", "µg", "spurenelement", None, 166, 167),
]

# Textmarker, die "kein Wert" bedeuten. Alles mit "<" wird gesondert behandelt.
MARKER_OHNE_WERT = {"-", "--", "–", "—", "tr", "spuren", "n.b.", "n.a.", "k.a."}

SCHEMA = """
CREATE TABLE IF NOT EXISTS lebensmittel (
    lebensmittel_id INTEGER PRIMARY KEY,
    herkunft        TEXT NOT NULL,
    bls_schluessel  TEXT,
    bezeichnung     TEXT NOT NULL,
    basis_menge_g   REAL NOT NULL DEFAULT 100,
    hersteller      TEXT,
    archiviert      INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS naehrstoff (
    naehrstoff_id    INTEGER PRIMARY KEY,
    bls_spalte       TEXT NOT NULL,
    name             TEXT NOT NULL,
    einheit          TEXT NOT NULL,
    gruppe           TEXT NOT NULL,
    uebergeordnet_id INTEGER REFERENCES naehrstoff(naehrstoff_id)
);

CREATE TABLE IF NOT EXISTS naehrwert (
    lebensmittel_id INTEGER NOT NULL REFERENCES lebensmittel(lebensmittel_id),
    naehrstoff_id   INTEGER NOT NULL REFERENCES naehrstoff(naehrstoff_id),
    wert_je_100g    REAL NOT NULL,
    wert_herkunft   TEXT,
    PRIMARY KEY (lebensmittel_id, naehrstoff_id)
);

CREATE INDEX IF NOT EXISTS idx_lebensmittel_bls ON lebensmittel(bls_schluessel);
CREATE INDEX IF NOT EXISTS idx_naehrwert_naehrstoff ON naehrwert(naehrstoff_id);
"""


def pruefe_kopfzeile(kopf):
    """Prueft, dass hinter jeder Spaltennummer wirklich der erwartete Code steht."""
    fehler = []

    if len(kopf) != ERWARTETE_SPALTENZAHL:
        fehler.append(
            f"Die Datei hat {len(kopf)} Spalten, erwartet waren {ERWARTETE_SPALTENZAHL}."
        )

    def ueberschrift(nummer):
        if nummer > len(kopf):
            return None
        wert = kopf[nummer - 1]
        return str(wert).strip() if wert is not None else None

    kopf_code = ueberschrift(SPALTE_CODE)
    if not kopf_code or not kopf_code.upper().startswith("BLS"):
        fehler.append(f"Spalte {SPALTE_CODE} sollte der BLS-Code sein, steht aber: {kopf_code!r}")

    for _, code, name, _, _, _, spalte_wert, spalte_herkunft in NAEHRSTOFFE:
        text_wert = ueberschrift(spalte_wert)
        text_herkunft = ueberschrift(spalte_herkunft)

        if not text_wert or text_wert.split()[0] != code:
            fehler.append(
                f"{name}: Spalte {spalte_wert} sollte mit {code!r} beginnen, "
                f"Ueberschrift ist aber {text_wert!r}"
            )
        if not text_herkunft or text_herkunft.split()[0] != code:
            fehler.append(
                f"{name}: Spalte {spalte_herkunft} sollte mit {code!r} beginnen, "
                f"Ueberschrift ist aber {text_herkunft!r}"
            )
        elif "Datenherkunft" not in text_herkunft:
            fehler.append(
                f"{name}: Spalte {spalte_herkunft} sollte die Datenherkunft sein, "
                f"Ueberschrift ist aber {text_herkunft!r}"
            )

    if fehler:
        print("Abbruch. Der Spaltenaufbau der Quelldatei passt nicht:", file=sys.stderr)
        for zeile in fehler:
            print("  - " + zeile, file=sys.stderr)
        sys.exit(1)


def lies_wert(zelle):
    """Gibt (wert, grund) zurueck. Genau eines von beiden ist None.

    wert ist immer eine Zahl, niemals Text. grund benennt, warum uebersprungen wird.
    """
    if zelle is None:
        return None, "leer"
    if isinstance(zelle, bool):
        return None, "unbekannt"
    if isinstance(zelle, (int, float)):
        return float(zelle), None

    text = str(zelle).strip()
    if not text:
        return None, "leer"

    vergleich = text.lower().replace(" ", "")
    if vergleich in MARKER_OHNE_WERT:
        return None, "marker"
    if vergleich.startswith("<"):
        # <LOD, <LOQ, <LOD or <LOQ: unterhalb Nachweis- oder Bestimmungsgrenze
        return None, "unter_grenze"

    # Zahl, die als Text in der Zelle steht (auch mit Komma als Trennzeichen)
    try:
        return float(text.replace(",", ".")), None
    except ValueError:
        return None, "unbekannt"


def lies_text(zelle):
    if zelle is None:
        return None
    text = str(zelle).strip()
    return text or None


def bestehende_bls_daten(verbindung):
    try:
        return verbindung.execute(
            "SELECT COUNT(*) FROM lebensmittel WHERE herkunft = 'bls'"
        ).fetchone()[0]
    except sqlite3.OperationalError:
        return 0


def vorhandene_schluessel(verbindung):
    """bls_schluessel -> lebensmittel_id der bereits angelegten BLS-Lebensmittel."""
    try:
        zeilen = verbindung.execute(
            "SELECT bls_schluessel, lebensmittel_id FROM lebensmittel "
            "WHERE herkunft = 'bls' AND bls_schluessel IS NOT NULL"
        ).fetchall()
    except sqlite3.OperationalError:
        return {}
    return {schluessel: kennung for schluessel, kennung in zeilen}


def entferne_naehrwerte(verbindung, kennungen):
    """Löscht die Nährwerte der genannten Lebensmittel, nicht die Lebensmittel selbst.

    Auf naehrwert verweist nichts, auf lebensmittel dagegen schon: Positionen in
    mahlzeit_position. Deshalb bleiben die Zeilen in lebensmittel bestehen und
    behalten ihre lebensmittel_id, damit erfasste Mahlzeiten gültig bleiben.
    """
    for anfang in range(0, len(kennungen), 500):
        teil = kennungen[anfang : anfang + 500]
        platzhalter = ",".join("?" * len(teil))
        verbindung.execute(
            f"DELETE FROM naehrwert WHERE lebensmittel_id IN ({platzhalter})", teil
        )


def schreibe_naehrstoffe(verbindung):
    """Legt die Naehrstoffe an. Die Verweise folgen in einem zweiten Durchgang.

    Ein uebergeordneter Naehrstoff kann eine hoehere Kennung haben als sein
    Kind (NACL verweist auf NA). Erst alle Zeilen schreiben, dann die Verweise
    setzen, damit die Reihenfolge in der Liste keine Rolle spielt.
    """
    nach_code = {code: nid for nid, code, *_ in NAEHRSTOFFE}
    for nid, code, name, einheit, gruppe, _, _, _ in NAEHRSTOFFE:
        verbindung.execute(
            "INSERT OR REPLACE INTO naehrstoff "
            "(naehrstoff_id, bls_spalte, name, einheit, gruppe, uebergeordnet_id) "
            "VALUES (?, ?, ?, ?, ?, NULL)",
            (nid, code, name, einheit, gruppe),
        )
    for nid, code, _, _, _, uebergeordnet, _, _ in NAEHRSTOFFE:
        if uebergeordnet is None:
            continue
        verbindung.execute(
            "UPDATE naehrstoff SET uebergeordnet_id = ? WHERE naehrstoff_id = ?",
            (nach_code[uebergeordnet], nid),
        )
    return nach_code


def importiere(quelle, datenbank, ersetzen):
    if not quelle.exists():
        sys.exit(f"Quelldatei nicht gefunden: {quelle}")

    print(f"Quelle:    {quelle}")
    print(f"Datenbank: {datenbank}")

    verbindung = sqlite3.connect(datenbank)
    verbindung.execute("PRAGMA foreign_keys = ON")
    verbindung.executescript(SCHEMA)

    vorhanden = bestehende_bls_daten(verbindung)
    if vorhanden and not ersetzen:
        verbindung.close()
        sys.exit(
            f"Abbruch. In {datenbank.name} stehen bereits {vorhanden} BLS-Lebensmittel.\n"
            "Erneut einlesen mit: python importe/import_bls.py --ersetzen\n"
            "Dabei werden vorhandene Eintraege anhand ihres BLS-Schluessels aktualisiert, "
            "nicht geloescht; erfasste Mahlzeiten bleiben gueltig."
        )
    bekannt = vorhandene_schluessel(verbindung)
    if vorhanden:
        print(f"Aktualisiere {vorhanden} vorhandene BLS-Lebensmittel anhand ihres Schluessels.")
        entferne_naehrwerte(verbindung, list(bekannt.values()))

    schreibe_naehrstoffe(verbindung)

    mappe = openpyxl.load_workbook(quelle, read_only=True, data_only=True)
    blatt = mappe[mappe.sheetnames[0]]
    zeilen = blatt.iter_rows(min_row=1, values_only=True)

    kopf = next(zeilen)
    pruefe_kopfzeile(kopf)

    anzahl_lebensmittel = 0
    neu = 0
    aktualisiert = 0
    gesehen = set()
    zeilen_gelesen = 0
    ohne_bezeichnung = 0
    mit_wert = Counter()
    uebersprungen = Counter()
    gruende = {code: Counter() for _, code, *_ in NAEHRSTOFFE}
    unbekannte_marker = {code: Counter() for _, code, *_ in NAEHRSTOFFE}

    for zeile in zeilen:
        schluessel = lies_text(zeile[SPALTE_CODE - 1])
        if schluessel is None:
            continue  # Leerzeilen am Dateiende
        zeilen_gelesen += 1

        bezeichnung = lies_text(zeile[SPALTE_BEZEICHNUNG - 1])
        if bezeichnung is None:
            ohne_bezeichnung += 1
            bezeichnung = schluessel

        gesehen.add(schluessel)
        lebensmittel_id = bekannt.get(schluessel)
        if lebensmittel_id is None:
            cursor = verbindung.execute(
                "INSERT INTO lebensmittel "
                "(herkunft, bls_schluessel, bezeichnung, basis_menge_g, hersteller, archiviert) "
                "VALUES ('bls', ?, ?, 100, NULL, 0)",
                (schluessel, bezeichnung),
            )
            lebensmittel_id = cursor.lastrowid
            bekannt[schluessel] = lebensmittel_id
            neu += 1
        else:
            # Kennung bleibt erhalten, damit Verweise aus mahlzeit_position gültig
            # bleiben. Ein früher archivierter Eintrag ist wieder aktuell.
            verbindung.execute(
                "UPDATE lebensmittel SET bezeichnung = ?, archiviert = 0 "
                "WHERE lebensmittel_id = ?",
                (bezeichnung, lebensmittel_id),
            )
            aktualisiert += 1
        anzahl_lebensmittel += 1

        for nid, code, _, _, _, _, spalte_wert, spalte_herkunft in NAEHRSTOFFE:
            wert, grund = lies_wert(zeile[spalte_wert - 1])
            if wert is None:
                uebersprungen[code] += 1
                gruende[code][grund] += 1
                if grund == "unbekannt":
                    unbekannte_marker[code][str(zeile[spalte_wert - 1]).strip()] += 1
                continue

            verbindung.execute(
                "INSERT INTO naehrwert "
                "(lebensmittel_id, naehrstoff_id, wert_je_100g, wert_herkunft) "
                "VALUES (?, ?, ?, ?)",
                (lebensmittel_id, nid, wert, lies_text(zeile[spalte_herkunft - 1])),
            )
            mit_wert[code] += 1

    # Was in dieser Ausgabe fehlt, wird archiviert statt gelöscht, damit alte
    # Mahlzeiten nachvollziehbar bleiben.
    verschwunden = [
        kennung for schluessel, kennung in bekannt.items() if schluessel not in gesehen
    ]
    for anfang in range(0, len(verschwunden), 500):
        teil = verschwunden[anfang : anfang + 500]
        platzhalter = ",".join("?" * len(teil))
        verbindung.execute(
            f"UPDATE lebensmittel SET archiviert = 1 WHERE lebensmittel_id IN ({platzhalter})",
            teil,
        )

    verbindung.commit()
    mappe.close()

    zeilen_naehrwert = verbindung.execute("SELECT COUNT(*) FROM naehrwert").fetchone()[0]
    verbindung.close()

    bericht(
        zeilen_gelesen,
        anzahl_lebensmittel,
        ohne_bezeichnung,
        zeilen_naehrwert,
        mit_wert,
        uebersprungen,
        gruende,
        unbekannte_marker,
        neu=neu,
        aktualisiert=aktualisiert,
        archiviert=len(verschwunden),
    )


def bericht(
    zeilen_gelesen,
    anzahl_lebensmittel,
    ohne_bezeichnung,
    zeilen_naehrwert,
    mit_wert,
    uebersprungen,
    gruende,
    unbekannte_marker,
    neu=0,
    aktualisiert=0,
    archiviert=0,
):
    print()
    print("Import abgeschlossen")
    print(f"  Datenzeilen gelesen:       {zeilen_gelesen}")
    print(f"  Lebensmittel verarbeitet:  {anzahl_lebensmittel}")
    print(f"    davon neu angelegt:      {neu}")
    print(f"    davon aktualisiert:      {aktualisiert}")
    print(f"  archiviert (nicht mehr in der Quelle): {archiviert}")
    print(f"  Zeilen in naehrwert:       {zeilen_naehrwert}")
    if ohne_bezeichnung:
        print(f"  Ohne Bezeichnung (Schluessel eingesetzt): {ohne_bezeichnung}")

    print()
    kopf = f"  {'Naehrstoff':<36}{'Code':<10}{'mit Wert':>10}{'uebersprungen':>15}   Gruende"
    print(kopf)
    print("  " + "-" * (len(kopf) - 2))
    for _, code, name, einheit, _, _, _, _ in NAEHRSTOFFE:
        aufschluesselung = ", ".join(
            f"{grund_text(g)}: {n}" for g, n in sorted(gruende[code].items())
        )
        print(
            f"  {name + ' [' + einheit + ']':<36}{code:<10}"
            f"{mit_wert[code]:>10}{uebersprungen[code]:>15}   {aufschluesselung or '-'}"
        )

    offene_marker = {c: m for c, m in unbekannte_marker.items() if m}
    if offene_marker:
        print()
        print("  ACHTUNG: nicht eingeplanter Zellinhalt, wurde uebersprungen:")
        for code, zaehler in offene_marker.items():
            for text, n in sorted(zaehler.items()):
                print(f"    {code}: {text!r} ({n}x)")


def grund_text(grund):
    return {
        "marker": "Strich/TR",
        "unter_grenze": "unter Nachweisgrenze",
        "leer": "leer",
        "unbekannt": "unbekannter Text",
    }.get(grund, grund)


def main():
    parser = argparse.ArgumentParser(description="BLS 4.0 nach tracker.db importieren")
    parser.add_argument("--quelle", type=Path, default=QUELLE, help="Pfad zur BLS-Excel-Datei")
    parser.add_argument("--db", type=Path, default=DATENBANK, help="Pfad zur SQLite-Datenbank")
    parser.add_argument(
        "--ersetzen",
        action="store_true",
        help="vorhandene Eintraege anhand ihres BLS-Schluessels aktualisieren",
    )
    argumente = parser.parse_args()
    importiere(argumente.quelle, argumente.db, argumente.ersetzen)


if __name__ == "__main__":
    main()
