"""Import der DGE/ÖGE-Referenzwerte in tracker.db.

Nicht Teil der Anwendung. Liest daten/DGE-Referenzwerte.xlsx mit den Blaettern
"Referenzwerte" und "Fußnoten" und fuellt die Tabellen referenzwert und fussnote.

Es wird nichts geraten: Nur die vorgesehenen Wertformate werden gelesen, alles
andere fuehrt zum Abbruch mit Angabe der Zeile. Naehrstoffbezeichnungen, die
nicht auf einen BLS-Code abgebildet sind, werden uebersprungen und am Ende
namentlich gemeldet.

Aufruf aus dem Projektordner:
    python importe/import_referenzwerte.py
"""

import argparse
import re
import sqlite3
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl fehlt. Installieren mit: pip install openpyxl")


# Das Skript liegt in importe/, die Datenbank und die Datenordner eine Ebene darueber.
BASIS = Path(__file__).resolve().parent.parent
QUELLE = BASIS / "daten" / "DGE-Referenzwerte.xlsx"
DATENBANK = BASIS / "tracker.db"

BLATT_WERTE = "Referenzwerte"
BLATT_FUSSNOTEN = "Fußnoten"

ERWARTETE_SPALTEN = (
    "Bevölkerungsgruppe", "Geschlecht", "Nährstoff", "Referenzwert",
    "Einheit", "Bemerkung", "Kategorie", "Fussnoten",
)

QUELLENANGABE = "DGE/ÖGE-Referenzwerte für die Nährstoffzufuhr"
STAND = "3. Auflage, 1. Ausgabe 2025"

# Bezeichnung der Quelle -> BLS-Code. Alles, was hier fehlt, wird uebersprungen.
# Von den drei Zinkstufen wird nur die mittlere genommen, sie entspricht einer
# gemischten Kost; fuer die Kinderaltersgruppen liefert die Quelle Zink ohne
# Staffelung.
NAEHRSTOFF_ZUORDNUNG = {
    "Vitamin A": "VITAA",
    "Vitamin D": "VITD",
    "Vitamin E": "VITE",
    "Vitamin K": "VITK",
    "Thiamin": "THIA",
    "Riboflavin": "RIBF",
    "Niacin": "NIAEQ",
    "Vitamin B6": "VITB6",
    "Folat": "FOL",
    "Pantothensäure": "PANTAC",
    "Biotin": "BIOT",
    "Vitamin B12 (Cobalamine)": "VITB12",
    "Vitamin C": "VITC",
    "Natrium": "NA",
    "Chlorid": "CLD",
    "Kalium": "K",
    "Calcium": "CA",
    "Phosphor": "P",
    "Magnesium": "MG",
    "Eisen": "FE",
    "Jod": "ID",
    "Fluorid": "FD",
    "Zink": "ZN",
    "Zink bei mittlerer Phytatzufuhr": "ZN",
    "Kupfer": "CU",
    "Mangan": "MN",
    "Chrom": "CR",
    "Molybdän": "MO",
    "Protein": "PROT625",
    "Gesamtfett": "FAT",
    "Gesättigte Fettsäuren": "FASAT",
    "Kohlenhydrate": "CHO",
}

KATEGORIEN = {
    "Empfohlene Zufuhr": "empfehlung",
    "Schätzwert": "schaetzwert",
    "Richtwert": "richtwert",
}

# Die Einheit der Quelle bestimmt den Bezug des Wertes.
BEZUG_JE_EINHEIT = {
    "mg/Tag": "absolut",
    "µg/Tag": "absolut",
    "µg-RAE/Tag": "absolut",
    "g/Tag": "absolut",
    "g/kg KG/Tag": "je_kg",
    "% der Energie": "prozent_energie",
}

GESCHLECHTER = {"Männlich": "m", "Weiblich": "w"}

# Einheit der Quelle -> Einheit, in der der Zahlenwert steht. Dient nur dem
# Abgleich mit naehrstoff.einheit, umgerechnet wird nichts.
MASSEINHEIT_JE_EINHEIT = {
    "mg/Tag": "mg",
    "µg/Tag": "µg",
    "µg-RAE/Tag": "µg",
    "g/Tag": "g",
    "g/kg KG/Tag": "g",
}

# Obergrenze fuer Altersgruppen der Form "65 Jahre und älter".
OFFENES_ENDE = 999

SCHEMA = """
CREATE TABLE IF NOT EXISTS referenzwert (
    referenzwert_id INTEGER PRIMARY KEY,
    naehrstoff_id   INTEGER NOT NULL REFERENCES naehrstoff(naehrstoff_id),
    geschlecht      TEXT NOT NULL,
    alter_von_jahre INTEGER NOT NULL,
    alter_bis_jahre INTEGER NOT NULL,
    art             TEXT,
    bezug           TEXT NOT NULL,
    wert            REAL,
    obergrenze      REAL,
    bemerkung       TEXT,
    fussnoten       TEXT,
    quelle          TEXT NOT NULL,
    stand           TEXT NOT NULL,
    UNIQUE (naehrstoff_id, geschlecht, alter_von_jahre, alter_bis_jahre)
);

CREATE TABLE IF NOT EXISTS fussnote (
    kuerzel      TEXT PRIMARY KEY,
    beschreibung TEXT NOT NULL
);
"""


class Abbruch(Exception):
    """Ein Feld passt in keines der vorgesehenen Muster."""


def text(zelle):
    if zelle is None:
        return None
    wert = str(zelle).strip()
    return wert or None


def zahl(rohtext: str, zeilennummer: int, feld: str) -> float:
    """Wandelt eine Zahl der Quelle um. Das Komma ist Dezimaltrennzeichen."""
    versuch = rohtext.strip().replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(versuch)
    except ValueError as fehler:
        raise Abbruch(f"Zeile {zeilennummer}: {feld} nicht lesbar: {rohtext!r}") from fehler


def lies_altersgruppe(rohtext: str, zeilennummer: int) -> tuple[int, int]:
    """25 bis unter 51 Jahre -> (25, 51); 65 Jahre und älter -> (65, 999).

    Untergrenze einschliesslich, Obergrenze ausschliesslich.
    """
    treffer = re.fullmatch(r"(\d+)\s*bis unter\s*(\d+)\s*Jahre", rohtext.strip())
    if treffer:
        return int(treffer.group(1)), int(treffer.group(2))
    treffer = re.fullmatch(r"(\d+)\s*Jahre und älter", rohtext.strip())
    if treffer:
        return int(treffer.group(1)), OFFENES_ENDE
    raise Abbruch(f"Zeile {zeilennummer}: Altersgruppe nicht lesbar: {rohtext!r}")


def lies_wert(zelle, zeilennummer: int) -> tuple[float | None, float | None, str | None] | None:
    """Gibt (wert, obergrenze, freitext) zurueck oder None fuer einen Strich.

    Sechs Formate der Quelle:
      Zahl            16 oder 0,8            -> wert
      Bereich         30-100                 -> wert und obergrenze
      Hoechstmenge    max. 10                -> nur obergrenze
      Mindestmenge    >=30 oder >4           -> wert
      Strich          -                      -> keine Zeile
      Freitext        beginnt mit einer Zahl -> erste Zahl als wert, Original in bemerkung
    """
    if isinstance(zelle, (int, float)) and not isinstance(zelle, bool):
        return float(zelle), None, None

    rohtext = text(zelle)
    if rohtext is None:
        raise Abbruch(f"Zeile {zeilennummer}: Referenzwert ist leer")

    if re.fullmatch(r"[-–—]+", rohtext):
        return None

    if re.fullmatch(r"[\d.,\s]+", rohtext):
        return zahl(rohtext, zeilennummer, "Referenzwert"), None, None

    treffer = re.fullmatch(r"([\d.,]+)\s*[-–]\s*([\d.,]+)", rohtext)
    if treffer:
        return (
            zahl(treffer.group(1), zeilennummer, "Untergrenze"),
            zahl(treffer.group(2), zeilennummer, "Obergrenze"),
            None,
        )

    treffer = re.fullmatch(r"max\.?\s*([\d.,]+)", rohtext, re.IGNORECASE)
    if treffer:
        return None, zahl(treffer.group(1), zeilennummer, "Hoechstmenge"), None

    treffer = re.fullmatch(r"[≥>]\s*([\d.,]+)", rohtext)
    if treffer:
        return zahl(treffer.group(1), zeilennummer, "Mindestmenge"), None, None

    # Freitext: nur zulaessig, wenn er mit einer Zahl beginnt. Diese Zahl ist der
    # Wert, der vollstaendige Text wird als Bemerkung uebernommen.
    treffer = re.match(r"^([\d.,]+)\s*\D", rohtext)
    if treffer:
        return zahl(treffer.group(1), zeilennummer, "Freitext"), None, rohtext

    raise Abbruch(f"Zeile {zeilennummer}: Referenzwert in keinem bekannten Format: {rohtext!r}")


def pruefe_kopfzeile(kopf, zeilennummer=1):
    vorhanden = [text(z) for z in kopf]
    fehlend = [name for name in ERWARTETE_SPALTEN if name not in vorhanden]
    if fehlend:
        print("Abbruch. Die Kopfzeile des Blattes passt nicht:", file=sys.stderr)
        print(f"  gefunden: {vorhanden}", file=sys.stderr)
        print(f"  fehlt:    {fehlend}", file=sys.stderr)
        sys.exit(1)


def naehrstoff_stammdaten(verbindung) -> dict[str, sqlite3.Row]:
    try:
        zeilen = verbindung.execute(
            "SELECT bls_spalte, naehrstoff_id, einheit FROM naehrstoff"
        ).fetchall()
    except sqlite3.OperationalError:
        sys.exit(
            "Die Tabelle naehrstoff fehlt. Bitte zuerst "
            "python importe/import_bls.py ausfuehren."
        )
    return {zeile["bls_spalte"]: zeile for zeile in zeilen}


def schreibe_fussnoten(verbindung, mappe) -> int:
    if BLATT_FUSSNOTEN not in mappe.sheetnames:
        sys.exit(f"Abbruch. Das Blatt {BLATT_FUSSNOTEN!r} fehlt in der Quelldatei.")
    anzahl = 0
    for zeile in mappe[BLATT_FUSSNOTEN].iter_rows(min_row=2, values_only=True):
        kuerzel, beschreibung = text(zeile[0]), text(zeile[1]) if len(zeile) > 1 else None
        if not kuerzel or not beschreibung:
            continue
        verbindung.execute(
            "INSERT INTO fussnote (kuerzel, beschreibung) VALUES (?, ?) "
            "ON CONFLICT (kuerzel) DO UPDATE SET beschreibung = excluded.beschreibung",
            (kuerzel, beschreibung),
        )
        anzahl += 1
    return anzahl


def importiere(quelle: Path, datenbank: Path) -> None:
    if not quelle.exists():
        sys.exit(f"Quelldatei nicht gefunden: {quelle}")

    print(f"Quelle:    {quelle}")
    print(f"Datenbank: {datenbank}")

    verbindung = sqlite3.connect(datenbank)
    verbindung.row_factory = sqlite3.Row
    verbindung.execute("PRAGMA foreign_keys = ON")
    verbindung.executescript(SCHEMA)

    stammdaten = naehrstoff_stammdaten(verbindung)
    kennungen = {code: zeile["naehrstoff_id"] for code, zeile in stammdaten.items()}
    mappe = openpyxl.load_workbook(quelle, data_only=True)
    if BLATT_WERTE not in mappe.sheetnames:
        sys.exit(f"Abbruch. Das Blatt {BLATT_WERTE!r} fehlt in der Quelldatei.")

    anzahl_fussnoten = schreibe_fussnoten(verbindung, mappe)

    blatt = mappe[BLATT_WERTE]
    zeilen = blatt.iter_rows(values_only=True)
    kopf = [text(z) for z in next(zeilen)]
    pruefe_kopfzeile(kopf)
    spalte = {name: kopf.index(name) for name in ERWARTETE_SPALTEN}

    gelesen = neu = aktualisiert = ohne_wert = 0
    je_kategorie = Counter()
    je_gruppe = Counter()
    uebersprungen = Counter()
    freitexte = []
    fehlende_codes = set()
    einheitenkonflikte = {}
    gesehen = set()

    try:
        for zeilennummer, roh in enumerate(zeilen, start=2):
            if not any(z is not None for z in roh):
                continue
            gelesen += 1

            bezeichnung = text(roh[spalte["Nährstoff"]])
            if bezeichnung is None:
                raise Abbruch(f"Zeile {zeilennummer}: Nährstoff fehlt")
            code = NAEHRSTOFF_ZUORDNUNG.get(bezeichnung)
            if code is None:
                uebersprungen[bezeichnung] += 1
                continue
            if code not in kennungen:
                fehlende_codes.add(f"{bezeichnung} -> {code}")
                continue

            alter_von, alter_bis = lies_altersgruppe(
                text(roh[spalte["Bevölkerungsgruppe"]]) or "", zeilennummer
            )
            geschlecht_roh = text(roh[spalte["Geschlecht"]])
            if geschlecht_roh not in GESCHLECHTER:
                raise Abbruch(f"Zeile {zeilennummer}: Geschlecht unbekannt: {geschlecht_roh!r}")
            geschlecht = GESCHLECHTER[geschlecht_roh]

            ergebnis = lies_wert(roh[spalte["Referenzwert"]], zeilennummer)
            if ergebnis is None:
                ohne_wert += 1
                continue
            wert, obergrenze, freitext = ergebnis

            einheit = text(roh[spalte["Einheit"]])
            bezug = BEZUG_JE_EINHEIT.get(einheit)
            if bezug is None:
                raise Abbruch(
                    f"Zeile {zeilennummer}: Einheit ohne Bezug: {einheit!r} ({bezeichnung})"
                )

            # Die Zahl bleibt so stehen, wie die Quelle sie angibt. Weicht die
            # Einheit von der in naehrstoff ab, wird das am Ende gemeldet: ein
            # spaeterer Vergleich muesste sonst um Faktor 1000 danebenliegen.
            masseinheit = MASSEINHEIT_JE_EINHEIT.get(einheit)
            if bezug == "absolut" and masseinheit and masseinheit != stammdaten[code]["einheit"]:
                einheitenkonflikte[code] = (bezeichnung, masseinheit, stammdaten[code]["einheit"])

            kategorie_roh = text(roh[spalte["Kategorie"]])
            if kategorie_roh is None:
                art = None
            elif kategorie_roh in KATEGORIEN:
                art = KATEGORIEN[kategorie_roh]
            else:
                raise Abbruch(f"Zeile {zeilennummer}: Kategorie unbekannt: {kategorie_roh!r}")

            bemerkung = text(roh[spalte["Bemerkung"]])
            if freitext:
                freitexte.append(f"{bezeichnung} {alter_von}-{alter_bis} {geschlecht}: {freitext}")
                bemerkung = freitext if bemerkung is None else f"{freitext} | {bemerkung}"
            fussnoten = text(roh[spalte["Fussnoten"]])

            schluessel = (kennungen[code], geschlecht, alter_von, alter_bis)
            if schluessel in gesehen:
                raise Abbruch(
                    f"Zeile {zeilennummer}: {bezeichnung} ({code}) kommt fuer diese "
                    f"Altersgruppe und dieses Geschlecht doppelt vor"
                )
            gesehen.add(schluessel)

            vorhanden = verbindung.execute(
                "SELECT 1 FROM referenzwert WHERE naehrstoff_id = ? AND geschlecht = ? "
                "AND alter_von_jahre = ? AND alter_bis_jahre = ?",
                schluessel,
            ).fetchone()
            verbindung.execute(
                "INSERT INTO referenzwert (naehrstoff_id, geschlecht, alter_von_jahre, "
                "alter_bis_jahre, art, bezug, wert, obergrenze, bemerkung, fussnoten, "
                "quelle, stand) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
                "ON CONFLICT (naehrstoff_id, geschlecht, alter_von_jahre, alter_bis_jahre) "
                "DO UPDATE SET art = excluded.art, bezug = excluded.bezug, "
                "wert = excluded.wert, obergrenze = excluded.obergrenze, "
                "bemerkung = excluded.bemerkung, fussnoten = excluded.fussnoten, "
                "quelle = excluded.quelle, stand = excluded.stand",
                (*schluessel, art, bezug, wert, obergrenze, bemerkung, fussnoten,
                 QUELLENANGABE, STAND),
            )
            if vorhanden:
                aktualisiert += 1
            else:
                neu += 1
            je_kategorie[art or "(ohne Kategorie)"] += 1
            je_gruppe[(alter_von, alter_bis, geschlecht)] += 1
    except Abbruch as fehler:
        verbindung.rollback()
        verbindung.close()
        mappe.close()
        print(f"Abbruch. {fehler}", file=sys.stderr)
        print("Es wurde nichts geschrieben.", file=sys.stderr)
        sys.exit(1)

    verbindung.commit()
    mappe.close()
    gesamt = verbindung.execute("SELECT COUNT(*) FROM referenzwert").fetchone()[0]
    verbindung.close()

    bericht(gelesen, neu, aktualisiert, ohne_wert, gesamt, anzahl_fussnoten,
            je_kategorie, je_gruppe, uebersprungen, freitexte, fehlende_codes,
            einheitenkonflikte)


def bericht(gelesen, neu, aktualisiert, ohne_wert, gesamt, anzahl_fussnoten,
            je_kategorie, je_gruppe, uebersprungen, freitexte, fehlende_codes,
            einheitenkonflikte):
    print()
    print("Import abgeschlossen")
    print(f"  Zeilen gelesen:             {gelesen}")
    print(f"  Referenzwerte neu:          {neu}")
    print(f"  Referenzwerte aktualisiert: {aktualisiert}")
    print(f"  ohne Wert (Strich):         {ohne_wert}")
    print(f"  Referenzwerte insgesamt:    {gesamt}")
    print(f"  Fussnoten:                  {anzahl_fussnoten}")

    print()
    print("  Je Kategorie:")
    for art, anzahl in sorted(je_kategorie.items()):
        print(f"    {art:<20} {anzahl}")

    print()
    print("  Je Altersgruppe und Geschlecht:")
    for (von, bis, geschlecht), anzahl in sorted(je_gruppe.items()):
        bereich = f"{von} bis unter {bis} Jahre" if bis != OFFENES_ENDE else f"ab {von} Jahren"
        print(f"    {bereich:<26} {geschlecht}   {anzahl} Werte")

    print()
    print(f"  Uebersprungene Naehrstoffbezeichnungen ({len(uebersprungen)}):")
    for bezeichnung, anzahl in sorted(uebersprungen.items()):
        print(f"    {bezeichnung} ({anzahl}x)")

    if freitexte:
        print()
        print("  Als Freitext gelesen, Original steht in bemerkung:")
        for eintrag in freitexte:
            print(f"    {eintrag}")

    if einheitenkonflikte:
        print()
        print("  ACHTUNG: Die Quelle gibt diese Werte in einer anderen Einheit an als")
        print("  naehrstoff.einheit. Die Zahlen stehen unveraendert in der Tabelle, ein")
        print("  Vergleich mit der Aufnahme muss die Einheit umrechnen:")
        for code, (bezeichnung, quelle_einheit, bls_einheit) in sorted(einheitenkonflikte.items()):
            print(f"    {code:<8} {bezeichnung:<24} Quelle {quelle_einheit}, naehrstoff {bls_einheit}")

    if fehlende_codes:
        print()
        print("  ACHTUNG: Zuordnung zeigt auf Codes, die es in naehrstoff nicht gibt:",
              file=sys.stderr)
        for eintrag in sorted(fehlende_codes):
            print(f"    {eintrag}", file=sys.stderr)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="DGE-Referenzwerte nach tracker.db importieren")
    parser.add_argument("--quelle", type=Path, default=QUELLE, help="Pfad zur Excel-Datei")
    parser.add_argument("--db", type=Path, default=DATENBANK, help="Pfad zur SQLite-Datenbank")
    argumente = parser.parse_args()
    importiere(argumente.quelle, argumente.db)


if __name__ == "__main__":
    main()
